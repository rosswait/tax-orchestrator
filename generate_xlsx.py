import argparse
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.comments import Comment
from datetime import datetime
import os
import json

def load_constants(target_year):
    """
    Search for tax constants in the constants/ directory.
    Uses the specified target_year if it exists, otherwise falls back to the highest available year.
    Returns (data_dict, logic_year)
    """
    base_dir = "constants"
    if not os.path.exists(base_dir):
        raise FileNotFoundError(f"Directory '{base_dir}' not found. Please ensure tax constants are present.")
    
    available_years = sorted([int(d) for d in os.listdir(base_dir) if d.isdigit()], reverse=True)
    if not available_years:
        raise FileNotFoundError("No year directories found in 'constants/'.")
    
    logic_year = target_year if target_year in available_years else available_years[0]
    
    files = {
        "fed_ord": "federal_ord.json",
        "fed_cg": "federal_cg.json",
        "ca": "ca_brackets.json",
        "surtaxes": "surtaxes.json"
    }
    
    data = {}
    years_used = {}
    for key, filename in files.items():
        found = False
        for yr in available_years:
            if yr > logic_year: continue
            path = os.path.join(base_dir, str(yr), filename)
            if os.path.exists(path):
                with open(path, "r") as f:
                    data[key] = json.load(f)
                years_used[key] = yr
                found = True
                break
        if not found:
            raise FileNotFoundError(f"Could not find {filename} in any year <= {logic_year}")
            
    return data, logic_year, years_used

def create_tax_workbook(status="Single", dependents=0, year=2026, fed_only=False, filename=None):
    wb = Workbook()
    
    # Define common format strings for Google Sheets parsing
    FORMAT_CURRENCY = '"$"#,##0.00'
    FORMAT_DATE = 'mm/dd/yyyy'
    FORMAT_PERCENT = '0.00%'

    # --- 0. Load Tax Logic Constants ---
    try:
        constants, LOGIC_YEAR, YEARS_USED = load_constants(year)
        fed_ord_brackets = constants["fed_ord"]
        fed_cg_brackets = constants["fed_cg"]
        ca_brackets = constants["ca"]
        surtaxes = constants["surtaxes"]
        
        fed_ord_start = 3
        fed_ord_end = fed_ord_start + len(fed_ord_brackets) - 1
        
        fed_cg_start = fed_ord_end + 4
        fed_cg_end = fed_cg_start + len(fed_cg_brackets) - 1
        
        if not fed_only:
            ca_start = fed_cg_end + 4
            ca_end = ca_start + len(ca_brackets) - 1
            surtax_start = ca_end + 4
        else:
            surtax_start = fed_cg_end + 4
            
        surtax_end = surtax_start + len(surtaxes) - 1
        
    except Exception as e:
        print(f"Error loading constants: {e}")
        return

    # --- 1. Instructions Tab (First) ---
    ws_instr = wb.active; ws_instr.title = "Instructions"
    ws_instr.append(["⚠️ DISCLAIMER: For Informational Purposes Only"])
    ws_instr.append(["This tool is intended for informational and educational purposes only and does not constitute professional tax, legal, or accounting advice. You should consult with a qualified tax professional before making financial decisions or filing your returns. The author assumes no liability for errors, omissions, or penalties resulting from the use of this workbook."])
    ws_instr.append([""])
    ws_instr.append(["Quick Start Guide"])
    ws_instr.append(["Step 1: Enter your latest YTD paystub details in the 'Wage Snapshots' tab."])
    ws_instr.append(["Step 2: Enter your latest YTD brokerage totals in 'Investment Income Snapshots'."])
    ws_instr.append(["Step 3: (Optional) Enter your 'Prior Year Tax' liability in the Dashboard to enable Safe Harbor targets."])
    ws_instr.append(["Step 4: (Optional) Enter your estimated 'Itemized Deductions' if you expect to exceed the Standard Deduction."])
    ws_instr.append(["Step 5: Check the 'PAYMENT ACTION CENTER' on the Dashboard for immediate payment requirements."])
    ws_instr.append([""])
    ws_instr.append(["Important Notes"])
    ws_instr.append(["1. Investment Income: Dividends and Interest are automatically projected for the full year based on the current quarter. Capital Gains (Short/Long term) are treated as one-off events and are NOT projected."])
    ws_instr.append(["2. HSA (California): HSA contributions are automatically added back to CA state income as they are not deductible in California."])
    ws_instr.append(["3. Income Projections: The engine pro-rates your remaining annual income based on the days elapsed since Jan 1. You can adjust these projections by entering values into 'Manual Income Offset' or 'Future Income Weight' in Section 1.5 of the Dashboard."])
    ws_instr.append(["4. YTD Methodology: Always use Year-to-Date (YTD) totals from your statements. Overwrite existing rows as new statements arrive."])

    # --- 2. Dashboard Tab (Second Position) ---
    ws_ds = wb.create_sheet("Dashboard")
    ws_ds["A1"] = "Configuration"
    ws_ds["A2"] = f"Filing Status ({status})"; ws_ds["B2"] = status
    ws_ds["A3"] = "Dependents (Under 17)"; ws_ds["B3"] = dependents
    ws_ds["A4"] = "Prior Year Fed Tax (Safe Harbor)"; ws_ds["B4"] = 0
    if not fed_only: ws_ds["A5"] = "Prior Year CA Tax (Safe Harbor)"; ws_ds["B5"] = 0
    ws_ds["A6"] = "Estimated Fed Itemized Deduction"; ws_ds["B6"] = 0
    if not fed_only: ws_ds["A7"] = "Estimated CA Itemized Deduction"; ws_ds["B7"] = 0
    ws_ds["A8"] = "Estimated Tax Year"; ws_ds["B8"] = "=IF(AND(MONTH(B9)=1, DAY(B9)<=30), YEAR(B9)-1, YEAR(B9))"
    ws_ds["A9"] = "Filing (Current) Date"; ws_ds["B9"] = "=TODAY()"
    ws_ds["A10"] = "Inferred Estimated Tax Quarter"; ws_ds["B10"] = "=MIN(4, MAX(1, ROUNDUP((MONTH(B9)-1)/3, 0)))"

    ws_ds["A12"] = "Projection & Assumptions"
    ws_ds["A13"] = "Manual Income Offset ($)"; ws_ds["B13"] = 0
    ws_ds["A14"] = "Future Income Weight (%)"; ws_ds["B14"] = 1.0
    ws_ds["A15"] = "Remaining Year Wage Income"; ws_ds["B15"] = "=IF(MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\")=0, 0, (SUM('Wage Snapshots'!C:C) / MAX(1, MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\") - DATE(B8,1,1))) * (DATE(B8,12,31) - MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\")))"
    ws_ds["A16"] = "Remaining Year Deductions"; ws_ds["B16"] = "=IF(MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\")=0, 0, (SUM('Wage Snapshots'!D:E) / MAX(1, MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\") - DATE(B8,1,1))) * (DATE(B8,12,31) - MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\")))"
    ws_ds["A17"] = "Remaining Year Interest and Dividends"; ws_ds["B17"] = "=SUM('Investment Income Snapshots'!C:C) * ( (4 / B10) - 1 )"

    ws_ds["A19"] = "Consolidated Income Projection"
    ws_ds["A20"] = "Total Projected Wage Income"; ws_ds["B20"] = "=SUM('Wage Snapshots'!C:C) + (B15 * B14) + B13"
    ws_ds["A21"] = "Federal W-2 State Wages"; ws_ds["B21"] = "=B20 - SUM('Wage Snapshots'!D:E) - (B16 * B14)"
    if not fed_only: ws_ds["A22"] = "CA W-2 State Wages"; ws_ds["B22"] = "=B20 - SUM('Wage Snapshots'!D:D) - (SUM('Wage Snapshots'!D:D)/MAX(1, SUM('Wage Snapshots'!D:E))) * B16 * B14"
    ws_ds["A23"] = "Investment Ordinary (Div/Int + STG)"; ws_ds["B23"] = "=SUM('Investment Income Snapshots'!C:C) + B17 + SUM('Investment Income Snapshots'!D:D)"
    ws_ds["A24"] = "Investment Preferential (LTG Only)"; ws_ds["B24"] = "=SUM('Investment Income Snapshots'!E:E)"
    ws_ds["A25"] = "Total Projected Federal AGI"; ws_ds["B25"] = "=B21 + B23 + B24"
    if not fed_only: ws_ds["A26"] = "Total Projected CA AGI"; ws_ds["B26"] = "=B22 + B23 + B24"

    ws_ds["A28"] = "Federal Tax Calculation"
    ws_ds["A29"] = "Deduction Applied (Max Std/Item)"; ws_ds["B29"] = f"=MAX(XLOOKUP(B2, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}, 'Tax Constants'!F{fed_ord_start}:F{fed_ord_end}, 0), B6)"
    ws_ds["A30"] = "Ordinary Taxable Income"; ws_ds["B30"] = "=MAX(0, B25 - B29 - B24)"
    ws_ds["A31"] = "Ordinary Income Tax"; ws_ds["B31"] = f"=XLOOKUP(B30, FILTER('Tax Constants'!C{fed_ord_start}:C{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), FILTER('Tax Constants'!D{fed_ord_start}:D{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), 0, -1) + (B30 - XLOOKUP(B30, FILTER('Tax Constants'!C{fed_ord_start}:C{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), FILTER('Tax Constants'!C{fed_ord_start}:C{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), 0, -1)) * XLOOKUP(B30, FILTER('Tax Constants'!C{fed_ord_start}:C{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), FILTER('Tax Constants'!E{fed_ord_start}:E{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), 0, -1)"
    ws_ds["A32"] = "Capital Gains Tax"; ws_ds["B32"] = f"=IF(B24>0, B24 * XLOOKUP(B30+B24, FILTER('Tax Constants'!C{fed_cg_start}:C{fed_cg_end}, 'Tax Constants'!B{fed_cg_start}:B{fed_cg_end}=B2), FILTER('Tax Constants'!D{fed_cg_start}:D{fed_cg_end}, 'Tax Constants'!B{fed_cg_start}:B{fed_cg_end}=B2), 0, -1), 0)"

    if not fed_only:
        ws_ds["A34"] = "CA Tax Calculation"
        ws_ds["A35"] = "CA Deduction Applied"; ws_ds["B35"] = f"=MAX(XLOOKUP(B2, 'Tax Constants'!B{ca_start}:B{ca_end}, 'Tax Constants'!F{ca_start}:F{ca_end}, 0), B7)"
        ws_ds["A36"] = "CA Taxable Income"; ws_ds["B36"] = "=MAX(0, B26 - B35)"
        ws_ds["A37"] = "CA Regular Tax"; ws_ds["B37"] = f"=XLOOKUP(B36, FILTER('Tax Constants'!C{ca_start}:C{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), FILTER('Tax Constants'!D{ca_start}:D{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), 0, -1) + (B36 - XLOOKUP(B36, FILTER('Tax Constants'!C{ca_start}:C{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), FILTER('Tax Constants'!C{ca_start}:C{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), 0, -1)) * XLOOKUP(B36, FILTER('Tax Constants'!C{ca_start}:C{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), FILTER('Tax Constants'!E{ca_start}:E{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), 0, -1)"
        ws_ds["A38"] = "CA MH Surcharge (1%)"; ws_ds["B38"] = "=IF(B36 > 1000000, (B36 - 1000000) * 0.01, 0)"
        ws_ds["A39"] = "Total CA Liability"; ws_ds["B39"] = "=B37 + B38"

    ws_ds["A41"] = "Final Liability & Surtaxes"
    ws_ds["A42"] = "NIIT Threshold"; ws_ds["B42"] = f"=XLOOKUP(B2, 'Tax Constants'!B{surtax_start}:B{surtax_end}, 'Tax Constants'!C{surtax_start}:C{surtax_end}, 200000)"
    ws_ds["A43"] = "NIIT (Fed)"; ws_ds["B43"] = "=IF(B25 > B42, 0.038 * MIN(B25-B42, B23+B24), 0)"
    ws_ds["A44"] = "Addl Medicare Threshold"; ws_ds["B44"] = f"=XLOOKUP(B2, 'Tax Constants'!B{surtax_start}:B{surtax_end}, 'Tax Constants'!D{surtax_start}:D{surtax_end}, 200000)"
    ws_ds["A45"] = "Addl Medicare (Fed)"; ws_ds["B45"] = "=IF(B20 > B44, 0.009 * (B20-B44), 0)"
    ws_ds["A46"] = "CTC Phaseout Start"; ws_ds["B46"] = f"=XLOOKUP(B2, 'Tax Constants'!B{surtax_start}:B{surtax_end}, 'Tax Constants'!E{surtax_start}:E{surtax_end}, 200000)"
    ws_ds["A47"] = "Child Tax Credit"; ws_ds["B47"] = "=IF(B25 > B46, MAX(0, (B3*2000)-ROUNDUP((B25-B46)/1000, 0)*50), B3*2000)"
    ws_ds["A48"] = "Total Federal Liability"; ws_ds["B48"] = "=B31 + B32 + B43 + B45 - B47"

    ws_ds["D1"] = "Estimated Tax Payments Ledger"
    ws_ds["D2"] = "Date"; ws_ds["E2"] = "Calculated Estimate (Optional)"; ws_ds["F2"] = "Actual Payment Made (Required)"; ws_ds["G2"] = "Note"
    ws_ds["D3"]=f"=DATE(B8,4,15)"; ws_ds["G3"]="Fed Q1"
    ws_ds["D5"]=f"=DATE(B8,6,15)"; ws_ds["G5"]="Fed Q2"
    ws_ds["D7"]=f"=DATE(B8,9,15)"; ws_ds["G7"]="Fed Q3"
    ws_ds["D9"]=f"=DATE(B8+1,1,15)"; ws_ds["G9"]="Fed Q4"
    if not fed_only:
        ws_ds["D4"]=f"=DATE(B8,4,15)"; ws_ds["G4"]="CA Q1"
        ws_ds["D6"]=f"=DATE(B8,6,15)"; ws_ds["G6"]="CA Q2"
        ws_ds["D8"]=f"=DATE(B8,9,15)"; ws_ds["G8"]="CA Q3"
        ws_ds["D10"]=f"=DATE(B8+1,1,15)"; ws_ds["G10"]="CA Q4"

    ws_ds["A50"] = "Payment Requirements"
    ws_ds["A51"] = "Fed Target"; ws_ds["B51"] = "=IF(B4=0, B48 * 0.9, MIN(B48 * 0.9, B4 * 1.1))"
    ws_ds["A52"] = "Total Fed Payments YTD"; ws_ds["B52"] = "=SUM('Wage Snapshots'!F:F) + SUMIFS(F3:F10, G3:G10, \"Fed*\")"
    if not fed_only:
        ws_ds["A53"] = "CA Target"; ws_ds["B53"] = "=IF(OR(B5=0, IF(B2=\"MFS\", B26>=500000, B26>=1000000)), B39 * 0.9, MIN(B39 * 0.9, B5 * 1.1))"
        ws_ds["A54"] = "Total CA Payments YTD"; ws_ds["B54"] = "=SUM('Wage Snapshots'!G:G) + SUMIFS(F3:F10, G3:G10, \"CA*\")"
    
    ws_ds["A57"] = "FEDERAL PAYMENT SCHEDULE"
    for i, q in enumerate(["Q1 (Apr 15)", "Q2 (Jun 15)", "Q3 (Sep 15)", "Q4 (Jan 15)"], 1):
        ws_ds[f"A{57+i}"] = q; ws_ds[f"B{57+i}"] = f"=B51 * {0.25*i}"; ws_ds[f"C{57+i}"] = f"=MAX(0, B{57+i} - B52)"
    
    if not fed_only:
        ws_ds["A64"] = "STATE PAYMENT SCHEDULE"
        for i, q in enumerate(["Q1 (Apr 15)", "Q2 (Jun 15)", "Q3 (Sep 15)", "Q4 (Jan 15)"], 1):
            rate = [0.3, 0.7, 0.7, 1.0][i-1]
            ws_ds[f"A{64+i}"] = q; ws_ds[f"B{64+i}"] = f"=B53 * {rate}"; ws_ds[f"C{64+i}"] = f"=MAX(0, B{64+i} - B54)"

    # --- Data Validation ---
    dv_status = DataValidation(type="list", formula1='"Single,MFJ,MFS,HoH"', showErrorMessage=True)
    ws_ds.add_data_validation(dv_status); dv_status.add(ws_ds["B2"])

    # --- Right Side Status Panel (Action Center) ---
    ws_ds["I1"] = "PAYMENT ACTION CENTER"
    ws_ds["I2"] = "FED DUE NOW:"; ws_ds["J2"] = "=IFS(B9<=DATE(B8,4,15), C58, B9<=DATE(B8,6,15), C59, B9<=DATE(B8,9,15), C60, TRUE, C61)"
    ws_ds["I3"] = "BY DEADLINE:"; ws_ds["J3"] = "=IFS(B9<=DATE(B8,4,15), \"04/15/\"&B8, B9<=DATE(B8,6,15), \"06/15/\"&B8, B9<=DATE(B8,9,15), \"09/15/\"&B8, TRUE, \"01/15/\"&(B8+1))"
    ws_ds["I4"] = "Next FED Target:"; ws_ds["J4"] = "=IFS(B9<=DATE(B8,4,15), B58, B9<=DATE(B8,6,15), B59, B9<=DATE(B8,9,15), B60, TRUE, B61)"
    ws_ds["I5"] = "FED Payments YTD:"; ws_ds["J5"] = "=B52"
    ws_ds["I6"] = "FED Status:"; ws_ds["J6"] = "=IF(J5>=J4, \"✅ Met (Surplus: \" & TEXT(J5-J4, \"$#,##0\") & \")\", \"🔴 Shortfall (\" & TEXT(J4-J5, \"$#,##0\") & \")\")"
    
    if not fed_only:
        ws_ds["I8"] = "CA DUE NOW:"; ws_ds["J8"] = "=IFS(B9<=DATE(B8,4,15), C65, B9<=DATE(B8,6,15), C66, B9<=DATE(B8,9,15), C67, TRUE, C68)"
        ws_ds["I9"] = "BY DEADLINE:"; ws_ds["J9"] = "=IFS(B9<=DATE(B8,4,15), \"04/15/\"&B8, B9<=DATE(B8,6,15), \"06/15/\"&B8, B9<=DATE(B8,9,15), \"09/15/\"&B8, TRUE, \"01/15/\"&(B8+1))"
        ws_ds["I10"] = "Next CA Target:"; ws_ds["J10"] = "=IFS(B9<=DATE(B8,4,15), B65, B9<=DATE(B8,6,15), B66, B9<=DATE(B8,9,15), B67, TRUE, B68)"
        ws_ds["I11"] = "CA Payments YTD:"; ws_ds["J11"] = "=B54"
        ws_ds["I12"] = "CA Status:"; ws_ds["J12"] = "=IF(J11>=J10, \"✅ Met (Surplus: \" & TEXT(J11-J10, \"$#,##0\") & \")\", \"🔴 Shortfall (\" & TEXT(J10-J11, \"$#,##0\") & \")\")"
    
    ws_ds["I15"] = "TAX DIAGNOSTICS"
    ws_ds["I16"] = "Fed Target Method:"; ws_ds["J16"] = "=IF(B4=0, \"90% Forecast\", IF(B48*0.9 < B4*1.1, \"90% Forecast\", \"110% Safe Harbor\"))"
    if not fed_only: ws_ds["I17"] = "CA Target Method:"; ws_ds["J17"] = "=IF(IF(B2=\"MFS\", B26>=500000, B26>=1000000), \"90% Forecast (AGI Limit)\", IF(B5=0, \"90% Forecast\", IF(B39*0.9 < B5*1.1, \"90% Forecast\", \"110% Safe Harbor\")))"
    ws_ds["I18"] = "Effective Fed Rate:"; ws_ds["J18"] = "=B48 / MAX(1, B25)"
    if not fed_only: ws_ds["I19"] = "Effective CA Rate:"; ws_ds["J19"] = "=B39 / MAX(1, B25)"
    ws_ds["I20"] = "Marginal Fed Bracket:"; ws_ds["J20"] = f"=XLOOKUP(B30, FILTER('Tax Constants'!C{fed_ord_start}:C{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), FILTER('Tax Constants'!E{fed_ord_start}:E{fed_ord_end}, 'Tax Constants'!B{fed_ord_start}:B{fed_ord_end}=B2), 0, -1)"
    if not fed_only: ws_ds["I21"] = "Marginal CA Bracket:"; ws_ds["J21"] = f"=XLOOKUP(B36, FILTER('Tax Constants'!C{ca_start}:C{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), FILTER('Tax Constants'!E{ca_start}:E{ca_end}, 'Tax Constants'!B{ca_start}:B{ca_end}=B2), 0, -1)"
    ws_ds["I22"] = "Deduction Applied:"; ws_ds["J22"] = "=IF(B6>XLOOKUP(B2, 'Tax Constants'!B3:B30, 'Tax Constants'!F3:F30, 0), \"ITEMIZED\", \"STANDARD\")"
    ws_ds["I23"] = "Fed Assumptions Year:"; ws_ds["J23"] = YEARS_USED["fed_ord"]
    ws_ds["I24"] = "CA Assumptions Year:"; ws_ds["J24"] = YEARS_USED.get("ca", "N/A")
    ws_ds["I25"] = "Inferred Quarter:"; ws_ds["J25"] = "=B10"
    
    ws_ds["I26"] = "ACTIVE WARNINGS"
    ws_ds["I27"] = "Stale Snapshots:"; ws_ds["J27"] = "=IF(OR(MAX('Wage Snapshots'!B:B)=0, (B9 - MAX('Wage Snapshots'!B:B)) > 30), \"🔴 !!! 30+ DAYS OLD !!!\", \"OK\")"
    ws_ds["I28"] = "Prior Year Data:"; ws_ds["J28"] = "=IF(B4=0, \"🔴 WARNING: FED MISSING\", \"OK\")"
    if not fed_only: ws_ds["I29"] = "HSA Verification (CA):"; ws_ds["J29"] = "=IF(SUM('Wage Snapshots'!E:E)=0, \"✅ No HSA Detected\", IF(B22=B21, \"🔴 ERR: HSA NOT ADDED TO CA\", \"✅ HSA Corrected (CA)\"))"
    ws_ds["I30"] = "Fed Brackets Stale:"; ws_ds["J30"] = f"=IF(B8 > J23, \"⚠️ FED STALE: \"&B8&\" brackets missing, using {YEARS_USED['fed_ord']} instead\", \"OK\")"
    if not fed_only: ws_ds["I31"] = "CA Brackets Stale:"; ws_ds["J31"] = f"=IF(B8 > J24, \"⚠️ CA STALE: \"&B8&\" brackets missing, using {YEARS_USED.get('ca', 'N/A')} instead\", \"OK\")"
    ws_ds["I32"] = "Filing Date Validity:"; ws_ds["J32"] = "=IF(OR(B9 < DATE(B8,1,1), B9 > DATE(B8+1,1,30)), \"🔴 ERR: DATE OUT OF RANGE\", \"OK\")"
    ws_ds["I33"] = "HOH Dependents:"; ws_ds["J33"] = "=IF(AND(B2=\"HoH\", B3=0), \"⚠️ Head of Household with 0 dependents is unusual\", \"OK\")"
    ws_ds["I34"] = "Stale Income Data:"; ws_ds["J34"] = "=IF(AND(MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\")>0, ROUNDUP(MONTH(MINIFS('Wage Snapshots'!B:B, 'Wage Snapshots'!B:B, \">0\"))/3, 0) < B10), \"⚠️ WAGES FROM PRIOR QTR\", IF(OR(AND(B10>1, COUNTIF('Investment Income Snapshots'!A:A, \"*Q1*\")>0), AND(B10>2, COUNTIF('Investment Income Snapshots'!A:A, \"*Q2*\")>0), AND(B10>3, COUNTIF('Investment Income Snapshots'!A:A, \"*Q3*\")>0)), \"⚠️ INV. INCOME FROM PRIOR QTR\", \"OK\"))"

    # --- 5. Data & Constants Tabs ---
    ws_inv = wb.create_sheet("Investment Income Snapshots")
    ws_inv.append(["Quarter", "Broker", "Dividends & Interest", "Short-Term Gains", "Long-Term Gains"])
    ws_wage = wb.create_sheet("Wage Snapshots")
    ws_wage.append(["Employer / Source", "Date", "Gross W-2 Income (YTD)", "Pre-tax Deductions (YTD)", "HSA Contributions (YTD)", "Fed Tax Withheld (YTD)", "CA Tax Withheld (YTD)", "FICA/Med/SDI (YTD)"])
    ws_const = wb.create_sheet("Tax Constants")
    ws_const.append(["Table A: Federal Brackets (Ordinary Income)"])
    ws_const.append(["Year", "Status", "Bracket Floor", "Base Tax", "Marginal Rate", "Standard Deduction"])
    for row in fed_ord_brackets: ws_const.append([YEARS_USED["fed_ord"]] + row)
    ws_const.append([]); ws_const.append(["Table B: Federal Capital Gains Brackets"])
    ws_const.append(["Year", "Status", "Bracket Floor", "LTCG Rate"])
    for row in fed_cg_brackets: ws_const.append([YEARS_USED["fed_cg"]] + row)
    if not fed_only:
        ws_const.append([]); ws_const.append(["Table C: California FTB Brackets"])
        ws_const.append(["Year", "Status", "Bracket Floor", "Base Tax", "Marginal Rate", "Standard Deduction", "MH Surcharge Floor"])
        for row in ca_brackets: ws_const.append([YEARS_USED.get("ca", "N/A")] + row)
    ws_const.append([]); ws_const.append(["Table D: Surtaxes & Phaseouts"])
    ws_const.append(["Year", "Status", "NIIT Threshold", "Addl Medicare", "CTC Phaseout Start"])
    for row in surtaxes: ws_const.append([YEARS_USED["surtaxes"]] + row)

    # --- 6. Parsing Instructions for Agents (Last) ---
    ws_ai = wb.create_sheet("Parsing Instructions for Agents")
    ws_ai.append(["PROMPT INSTRUCTIONS"])
    ws_ai.append(["Copy and paste the text below into an AI (like Gemini, ChatGPT, or Claude) along with your paystubs or brokerage statements to automatically extract the data for this workbook."])
    ws_ai.append([""])
    
    instr_path = "parsing_agent_instructions.md"
    if os.path.exists(instr_path):
        with open(instr_path, "r") as f:
            full_text = f.read()
        
        ws_ai["A4"] = full_text
        ws_ai["A4"].alignment = Alignment(wrap_text=True, vertical="top")
        ws_ai.column_dimensions["A"].width = 100
        ws_ai.row_dimensions[4].height = 600
    else:
        ws_ai.append(["[Error: parsing_agent_instructions.md not found. Please ensure the file exists in the current directory.]"])

    # --- Extensive Annotations ---
    ANN = {
        "D1": "Record your quarterly estimated tax payments made directly to the IRS or FTB here.",
        "F2": "ENTER PAYMENTS HERE. Required for DUE NOW calculations.",
        "B4": "Enter total tax from last year's Federal Form 1040.",
        "B6": "Populate with last year's itemized deductions unless your situation changed. Defaults to Standard Deduction if empty or lower than the threshold.",
        "B9": "The baseline date used for deadlines and current quarter inference (Filing Date).",
        "B10": "Automatically calculated based on the Filing Date using a 30-day buffer. This drives your Interest/Dividend projections.",
        "B13": "Manual Offset for one-off bonuses/non-recurring income.",
        "B14": "1.0 = normal earnings weight for projected future income.",
        "B17": "Projects Dividends/Interest for the remainder of the year based on the Inferred Quarter (B10).",
        "I1": "PAYMENT ACTION CENTER: High-level status board for current obligations.",
        "I15": "TAX DIAGNOSTICS: Real-time health check on rates and logic.",
        "I24": "Displays the quarter the logic is currently assuming for projections based on the Filing Date.",
        "I30": "Warns if Federal brackets are from prior years.",
        "I32": "Flags red if the Filing Date is before the Tax Year start or after the January buffer ends.",
        "I33": "Warns if Head of Household is selected with zero dependents (legal but rare).",
        "I34": "Flags if you have stale records from previous quarters which will artificially warp your pro-ration math."
    }
    if not fed_only:
        ANN.update({
            "B5": "Enter total tax from last year's California Form 540.",
            "B7": "Populate with last year's itemized deductions unless your situation changed. Defaults to Standard Deduction if empty or lower than the threshold.",
            "I29": "HSA Verification: Checks addition back to CA income.",
            "I31": "Warns if CA brackets are from prior years."
        })

    # --- Premium Formatting Engine ---
    st_sec = (Font(bold=True, size=11, color="FFFFFF"), PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid"))
    st_lbl = Font(bold=True); st_in = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    st_ac_bg = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    st_crit = Font(bold=True, color="FF0000"); st_calc = Font(bold=True)
    
    side = Side(border_style="thin", color="000000")
    st_border = Border(top=side, left=side, right=side, bottom=side)
    
    sec_k = ["Configuration", "Assumptions", "Projection", "Calculation", "Requirements", "Ledger", "SCHEDULE", "CENTER", "DIAGNOSTICS", "WARNINGS", "AI Parsing"]
    in_c = ["B2", "B3", "B4", "B5", "B6", "B7", "B9", "B13", "B14"]

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                coord = cell.coordinate; val = str(cell.value)
                if ws.title == "Dashboard" and coord in ANN:
                    cell.comment = Comment(ANN[coord], "Ross Wait")
                if ws.title == "Wage Snapshots" and coord == "A1":
                    cell.comment = Comment("YTD paystub data.", "Ross Wait")

                is_header = False
                if ws.title == "Dashboard" and any(k in val for k in sec_k):
                    is_header = True
                elif ws.title == "Instructions" and cell.row in [1, 4, 11]:
                    is_header = True
                elif ws.title == "Parsing Instructions for Agents" and cell.row in [1]:
                    is_header = True
                elif ws.title not in ["Dashboard", "Instructions", "Parsing Instructions for Agents"] and cell.row == 1:
                    is_header = True

                if is_header:
                    cell.font, cell.fill = st_sec
                elif (ws.title == "Dashboard" and cell.column in [1, 4, 9] and cell.row > 1):
                    cell.font = st_lbl
                
                if ws.title == "Dashboard":
                    if coord in in_c or (cell.column in [4,5,6,7] and 3 <= cell.row <= 10): cell.fill = st_in
                    if cell.column in [5, 6, 7] and 2 <= cell.row <= 10: cell.font = st_calc
                    if (9 <= cell.column <= 10 and 1 <= cell.row <= 13):
                        if cell.row == 1: cell.font, cell.fill = st_sec
                        else: cell.fill = st_ac_bg
                    if cell.coordinate in ["I2", "J2", "I8", "J8"]: cell.font = st_crit
                    
                    if cell.column == 2:
                        if cell.row in [8, 10]: cell.number_format = '0'
                        elif cell.row == 9: cell.number_format = FORMAT_DATE
                        elif cell.row == 14: cell.number_format = FORMAT_PERCENT
                        elif (4 <= cell.row <= 54) or (58 <= cell.row <= 75): cell.number_format = FORMAT_CURRENCY
                    if cell.column == 3 and (58 <= cell.row <= 75): cell.number_format = FORMAT_CURRENCY
                    if cell.column == 10:
                        if cell.row in [23, 24]: cell.number_format = '0'
                        elif cell.row in [2,4,5,8,10,11]: cell.number_format = FORMAT_CURRENCY
                        elif cell.row in [18,19]: cell.number_format = FORMAT_PERCENT
                        elif cell.row in [20,21]: cell.number_format = FORMAT_PERCENT
                    if cell.column == 4 and 3 <= cell.row <= 10: cell.number_format = FORMAT_DATE
                    if cell.column in [5,6] and 3 <= cell.row <= 10: cell.number_format = FORMAT_CURRENCY
                elif ws.title in ["Wage Snapshots", "Investment Income Snapshots"] and cell.row > 1:
                    if ws.title == "Wage Snapshots" and cell.column == 1: cell.number_format = FORMAT_DATE
                    else: cell.number_format = FORMAT_CURRENCY
        for col in ws.columns:
            l = max([len(str(c.value)) for c in col if c.value and not str(c.value).startswith('=')] + [10])
            ws.column_dimensions[col[0].column_letter].width = min(l + 2, 28)

    # Hide rows/columns in Fed-only mode
    if fed_only:
        for r in range(34, 40): ws_ds.row_dimensions[r].hidden = True # 34-39 CA Tax Calc
        for r in range(53, 55): ws_ds.row_dimensions[r].hidden = True # 53-54 CA Target
        for r in range(64, 69): ws_ds.row_dimensions[r].hidden = True # 64-68 CA Schedule
        ws_wage.column_dimensions['F'].hidden = True # Hide CA Tax Withheld

    if filename is None:
        if fed_only:
            filename = "Estimated_Tax_Calculator_FedOnly.xlsx"
        else:
            filename = "Estimated_Tax_Calculator.xlsx"

    wb.save(filename)
    print(f"Workbook polished and generated successfully: {filename}")

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--status", default="Single")
    parser.add_argument("--dependents", type=int, default=0)
    parser.add_argument("--year", type=int, default=datetime.now().year)
    parser.add_argument("--fed-only", action="store_true")
    parser.add_argument("--output", help="Custom output filename")
    
    args = parser.parse_args()
    create_tax_workbook(args.status, args.dependents, args.year, args.fed_only, args.output)